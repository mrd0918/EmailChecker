from openpyxl import load_workbook
import urllib.request
from urllib.request import Request, urlopen
import json
import time

wb = load_workbook('email.xlsx')
ws = wb['Sheet1']
checklist = []
for row in range(2,ws.max_row+1):
    for col in "B":
        cell_name = "{}{}".format(col,row)
        cell_value = ws[cell_name].value
        checklist.append(cell_value)
#print(checklist)
wb.create_sheet('result')
ws = wb['result']
#ws.active
ws['A1'].value = 'Email'
ws['B1'].value = 'Status'
ws['C1'].value = 'Count'
ws['D1'].value = 'Source'
ws['E1'].value = 'Verified'
ws['F1'].value = 'Date Leaked'
i = 2
for email in checklist:
    api = "https://haveibeenpwned.com/api/v2/breachedaccount/" + email
    request = Request(api, headers = {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/534.4 (KHTML, like Gecko) Chrome/6.0.481.0 Safari/534.4'}) 
    try:
        response = urllib.request.urlopen(request).read()
        content = json.loads(response.decode('utf-8'))
        le = len(content)
        for x in range(0, le):
            ws.cell(row=i, column=1).value = email
            ws.cell(row=i, column=2).value = "Found"
            ws.cell(row=i, column=3).value = le
            ws.cell(row=i, column=4).value = content[x]['Title']
            ws.cell(row=i, column=5).value = content[x]['IsVerified']
            ws.cell(row=i, column=6).value = content[x]['BreachDate']
            i = i +1
            wb.save('result.xlsx')
            time.sleep(2.5)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            ws.cell(row=i, column=1).value = email
            ws.cell(row=i, column=2).value = "Not Found"
            i = i + 1
            wb.save('result_pwn.xlsx')
            time.sleep(2.5)
        if e.code == 429:
            print("the rate limit has been exceeded")
        if e.code == 400:
            print("the account does not comply with an acceptable format")
        if e.code == 403:
            print("no user agent has been specified in the request")
wb.save('result.xlsx')

print('Completed')

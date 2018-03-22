# add for test
from openpyxl import load_workbook
import urllib.request
from urllib.request import Request, urlopen
import json
import datetime
import time
import random

wb = load_workbook('email.xlsx')
ws = wb['Sheet1']
checklist = []
for row in range(2,ws.max_row+1):
    for col in "B":
        cell_name = "{}{}".format(col,row)
        cell_value = ws[cell_name].value
        checklist.append(cell_value)
#print(checklist)
wb.create_sheet('result')
ws = wb['result']
#ws.active
ws['A1'].value = 'Email'
ws['B1'].value = 'Status'
ws['C1'].value = 'Count'
ws['D1'].value = 'Source'
ws['E1'].value = 'Verified'
ws['F1'].value = 'Date Leaked'
i = 2
for email in checklist:
    api = "https://hacked-emails.com/api?q=" + email
    user_agent = ['Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/534.4 (KHTML, like Gecko) Chrome/6.0.481.0 Safari/534.4',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.246',
                  'Mozilla/5.0 (X11; CrOS x86_64 8172.45.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.64 Safari/537.36',
                  'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_2) AppleWebKit/601.3.9 (KHTML, like Gecko) Version/9.0.2 Safari/601.3.9',
                  'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36',
                  'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1',
                  'Mozilla/5.0 (Linux; Android 7.0; SM-G892A Build/NRD90M; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/60.0.3112.107 Mobile Safari/537.36',
                  'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1'
                  'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.34 (KHTML, like Gecko) Version/11.0 Mobile/15A5341f Safari/604.1'
                  'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A5370a Safari/604.1'
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'
                  'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36'
                  'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko'
                  'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko'
                  'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko'
                  'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)'
                  ]
    #proxies = urllib.request.ProxyHandler({'http':'http://122.216.120.253:80',
    #                                       'http':'http://58.82.151.37:8080',
    #                                       'http':'http://50.233.137.34:80'})
    #opener = urllib.request.build_opener(proxies)
    #urllib.request.install_opener(opener)
    
    #response = urllib.request.urlopen(api)
    #content = response.read()
    
    request = Request(api, headers = {'User-Agent': random.choice(user_agent)})
    response = urllib.request.urlopen(request).read()
    content = json.loads(response.decode('utf-8'))
    if content['status'] == "notfound":
        ws.cell(row=i, column=1).value = email
        ws.cell(row=i, column=2).value = "Not Found"
        i = i + 1
        wb.save('result.xlsx')
        time.sleep(random.randint(5, 10))
        continue
    hit = int(content['results'])
    for x in range(0, hit):
        ws.cell(row=i, column=1).value = email
        ws.cell(row=i, column=2).value = "Found"
        ws.cell(row=i, column=3).value = hit
        ws.cell(row=i, column=4).value = content['data'][x]['title']
        ws.cell(row=i, column=5).value = content['data'][x]['verified']
        ws.cell(row=i, column=6).value = content['data'][x]['date_leaked']
        i = i + 1
        wb.save('result.xlsx')
        time.sleep(random.randint(5, 10))

wb.save('result.xlsx')
print('Completed')
